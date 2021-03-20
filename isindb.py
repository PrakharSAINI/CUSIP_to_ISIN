from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
import os
import pandas as pd
from openpyxl import load_workbook

#Set up browser (you have to dowload this chromedriver on https://chromedriver.chromium.org/downloads)
#Place the chromedriver file in the same folder as your project for this to work

chromedriver = "/Users/thibaultknobloch/Downloads/chromedriver"
os.environ["webdriver.chrome.driver"] = chromedriver
driver = webdriver.Chrome(chromedriver)

#Website I use to turn CUSIP into ISIN
url='https://www.isindb.com/fix-cusip-calculate-cusip-check-digit/'

#Load your cusip_codes, this works for cusip codes all in one column (will create a python list out of the column cells)
wb = load_workbook("cusip_excel.xlsx")  # Work Book
ws = wb.get_sheet_by_name('Sheet1')  # Work Sheet
column = ws['A']  # Column
data_list = [column[x].value for x in range(len(column))]

print(len(data_list))

#turn all the Cusips into strings to avoid float type errors (didn't work without this)
data_list2 = [str(i) for i in data_list]

fixed_cusip_list = []

#For each Cusip, go to website and fix the cusip (get the control digit)
for i in range(len(data_list2)):
    newKey = data_list2[i]
    driver.get(url)

    driver.find_element_by_id('userinput').send_keys(newKey)

    driver.find_element_by_id('fix').click()

    #cannot make this too short, otherwise some CUSIPs will not be retrived as they are not loaded
    driver.implicitly_wait(0.6)

    try:
        fixed_cusip = driver.find_element_by_css_selector('#result > p > strong')
        fixed_cusip_list.append(fixed_cusip.text)
    except NoSuchElementException:
        fixed_cusip_list.append(None)

#remove None values (when website gives CUSIP not valid error)
CUSIP_list = [x for x in fixed_cusip_list if x is not None]

print(len(CUSIP_list))

list = []

#url to convert fixed CUSIPs into ISIN
url='https://www.isindb.com/convert-cusip-to-isin/'

for b in range(len(CUSIP_list)):
    newKey = CUSIP_list[b]
    driver.get(url)

    driver.find_element_by_id('userinput').send_keys(newKey)

    driver.find_element_by_id('convert').click()
    driver.implicitly_wait(0.6)

    try:
        ISIN = driver.find_element_by_css_selector('#result > p > strong')
        list.append(ISIN.text)
    except NoSuchElementException:
        list.append(None)

ISIN_list = [x for x in list if x is not None]

print(len(ISIN_list))

#turn the list into excel file
ISIN_list = pd.DataFrame(ISIN_list, columns=["ISIN"])
ISIN_list.to_excel('ISIN_list.xlsx')

#DONE
print("done")


